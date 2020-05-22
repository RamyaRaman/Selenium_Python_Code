import openpyxl

path = 'G:\Python\TestData.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook.active #if only one sheet
#sheet = workbook.get_sheet_by_name("Sheet1") #if many sheet are there

rows = sheet.max_row  #11
cols = sheet.max_column  #4

print(rows)
print(cols)

#read the data from excel

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="    ")
    print()


